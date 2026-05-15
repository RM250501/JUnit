import os
import glob
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# ======= フォーマット用スタイル ========
bold = Font(bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
pink = PatternFill(start_color="FFD1E5FE", end_color="FFD1E5FE", fill_type="solid")
yellow = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
blue  = PatternFill(start_color="FFB7DEE8", end_color="FFB7DEE8", fill_type="solid")
gray  = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

# ======= 任意の抽出 & サンプルデータ自動埋め =======
def extract_conditions(code):
    # 本格if解析する場合はjavalang等を使う。今回はPDF見本用途サンプル自動生成
    return [
        'loginUser == null',
        'logic.getMyProvision()',
        'request.getRequestDispatcher("WEB-INF/jsp/Home.jsp").forward()',
        'throw new ServletException(e)'
    ]

def extract_code_snippet(code, max_lines=12):
    return "\n".join(code.splitlines()[:max_lines])

SRC_ROOT = '.'  # ルート配下全て

java_files = []
for folder, dirs, files in os.walk(SRC_ROOT):
    for file in files:
        if file.endswith('.java'):
            java_files.append(os.path.join(folder, file))

wb = openpyxl.Workbook()
wb.remove(wb.active)

for java_file in java_files:
    fname = os.path.basename(java_file)
    ws = wb.create_sheet(title=fname[:31])
    with open(java_file, encoding='utf-8') as f:
        code = f.read()

    # 1. タイトル行（結合＆ピンク帯）
    ws.merge_cells('A1:F1')
    ws['A1'] = "デシジョンテーブル"
    ws['A1'].font = bold
    ws['A1'].alignment = center
    ws['A1'].fill = pink

    # 2. ラベル部
    ws['A2'], ws['B2'], ws['C2'] = "チーム名", "チームC", ""
    ws['A3'], ws['B3'] = "試験対象プログラム名", fname
    ws['A4'], ws['B4'] = "簡単な説明", "サーブレットとしてHTTPリクエストを受け、入力値を解釈し業務ロジックを呼び出し、遷移先を決定します。"

    # 項目ごとの強調
    ws['A2'].font = bold
    ws['A3'].font = bold
    ws['A4'].font = bold
    for cell in ws['A2':'A4'][0]:
        cell.fill = pink

    # 3. テスト部ヘッダ（黄色帯＋結合）
    ws.merge_cells('A6:A7')
    ws['A6'] = "テストNo"
    ws['A6'].alignment = center
    ws['A6'].fill = yellow
    ws['A6'].font = bold
    ws['B6'] = "テスト担当者"
    ws['B6'].alignment = center
    ws['B6'].fill = yellow
    ws['B6'].font = bold
    ws.merge_cells('C6:E6')
    ws['C6'] = "テストケース"
    ws['C6'].alignment = center
    ws['C6'].fill = blue
    ws['C6'].font = bold
    ws['F6'] = "備考"
    ws['F6'].alignment = center
    ws['F6'].font = bold
    ws['F6'].fill = gray

    ws['B7'] = " "
    ws['C7'] = "GiveHome-01"
    ws['D7'] = "GiveHome-02"
    ws['E7'] = "GiveHome-03"

    for c in ['C7','D7','E7']:
        ws[c].alignment = center
        ws[c].fill = blue
        ws[c].font = bold

    # 4. 条件行
    conditions = extract_conditions(code)
    for row, cond in enumerate(conditions, start=8):
        ws[f'A{row}'] = ""
        ws[f'B{row}'] = cond
        ws[f'C{row}'] = "Y"
        ws[f'D{row}'] = "N"
        ws[f'E{row}'] = "X"
        ws[f'B{row}'].fill = yellow
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].alignment = center
            ws[f'{c}{row}'].border = border

    # 5. コード抜粋（下部）
    start_row = len(conditions) + 10
    ws[f'A{start_row}'] = "対象となるソースコード"
    ws[f'A{start_row}'].font = bold
    ws[f'A{start_row}'].fill = pink
    ws[f'A{start_row}'].alignment = left
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    code_txt = extract_code_snippet(code)
    ws[f'A{start_row+1}'] = code_txt
    ws[f'A{start_row+1}'].alignment = left
    ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+6, end_column=6)

    # 列幅
    for col, w in zip(['A','B','C','D','E','F'], [12,35,15,15,15,15]):
        ws.column_dimensions[col].width = w

wb.save('全ファイル_PDF体裁デシジョンテーブル.xlsx')
print("PDF見本体裁でExcel出力しました → 全ファイル_PDF体裁デシジョンテーブル.xlsx")