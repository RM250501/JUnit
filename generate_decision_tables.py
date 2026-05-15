import os
import glob
import openpyxl
import javalang

# Javaファイル探索ディレクトリ。リポジトリ構成によって適宜修正
SRC_ROOT = '.'  # 例: src/ 配下のすべての Java ファイルが対象

def extract_conditions_from_java(code):
    """if/switch などの条件式抽出"""
    conditions = []
    try:
        tree = javalang.parse.parse(code)
        for path, node in tree:
            if isinstance(node, javalang.tree.IfStatement) and node.condition:
                conditions.append('if: ' + str(node.condition))
            elif isinstance(node, javalang.tree.SwitchStatement) and node.expression:
                conditions.append('switch: ' + str(node.expression))
    except Exception:
        pass
    return conditions

# ファイル検出
java_files = glob.glob(os.path.join(SRC_ROOT, '**', '*.java'), recursive=True)
if not java_files:
    print(f"Javaファイルが '{SRC_ROOT}' 配下に見つかりません")
    exit()

wb = openpyxl.Workbook()
# 既存シート削除（openpyxl のデフォルトシート）
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

for java_file in java_files:
    fname = os.path.basename(java_file)
    sheet_name = fname[:31]  # シート名：31文字制限
    ws = wb.create_sheet(title=sheet_name)
    
    with open(java_file, encoding='utf-8') as f:
        code = f.read()
    
    # デシジョンテーブル上部情報
    ws['A1'] = 'デシジョンテーブル'
    ws['B2'] = '自動生成'
    ws['A3'] = 'ファイル名'
    ws['B3'] = fname
    ws['A4'] = 'パス'
    ws['B4'] = java_file
    ws['A5'] = '簡単な説明'
    ws['B5'] = 'このファイルの主要な分岐条件とコード抜粋'

    # テーブルヘッダ行
    ws.append(['テストNo', '条件', '分岐説明'])
    
    # 条件文自動抽出
    conditions = extract_conditions_from_java(code)
    if not conditions:
        ws.append([1, '（主要な条件文なし or パース不可）', ''])
    else:
        for i, cond in enumerate(conditions, 1):
            ws.append([i, cond, '自動抽出'])

    # 空行
    ws.append([])  
    ws.append(['--- コード抜粋 ---'])
    code_lines = code.splitlines()
    for idx, line in enumerate(code_lines[:20]):  # 先頭20行だけ抜粋
        ws.append([line])

    # 見易いよう列幅調整
    for col in ['A', 'B', 'C']:
        ws.column_dimensions[col].width = 50

# 保存
wb.save('全ファイル_デシジョンテーブル.xlsx')
print('Excel生成済み: 全ファイル_デシジョンテーブル.xlsx')